
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

def create_presence_template(worker_name, month, year):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.merge_cells('A1:D1')
    ws.merge_cells('E1:H1')
    ws['A1'] = f"NOME OPERAIO: {worker_name}"
    ws['E1'] = f"NOME OPERAIO: {worker_name}"
    ws['A2'] = f"MESE DI: {month:02d}/{year}"
    ws['E2'] = f"MESE DI: {month:02d}/{year}"

    headers = ["GIORNO", "CANTIERE", "N. ORE", "SIGLA DI CONTROLLO"] * 2
    ws.append(headers)

    for day in range(1, 32):
        row = [day, "", "", ""] + ([day+16, "", "", ""] if day+16 <= 31 else ["", "", "", ""])
        ws.append(row)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15

    filename = f"Presenze_{worker_name}_{month:02d}_{year}.xlsx"
    wb.save(filename)
    return filename
